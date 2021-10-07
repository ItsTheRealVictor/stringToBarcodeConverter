from barcode import Code128
from barcode.writer import ImageWriter
import openpyxl

wb = openpyxl.Workbook()
ws = wb.worksheets[0]

stringOne = 'Hank Hill'
stringTwo = 'Dale Gribble'
stringThree = 'John Redcorn'


convertedToBarcode1 = Code128(stringOne, writer=ImageWriter())
convertedToBarcode1.save(r"C:\Users\VD102541\Desktop\PDX_Archive_Barcode")

convertedToBarcode2 = Code128(stringTwo, writer=ImageWriter())
convertedToBarcode2.save(r'C:\Users\VD102541\Desktop\PDX_DeArchive_Barcode')

convertedToBarcode3 = Code128(stringThree, writer=ImageWriter())
convertedToBarcode3.save(r"C:\Users\VD102541\Desktop\PDX_ArchiveRecycle_Barcode")


ws['A1'] = stringOne
imageOne = openpyxl.drawing.image.Image(r"C:\Users\VD102541\Desktop\PDX_Archive_Barcode.png")
imageOne.anchor = 'E1'
ws.add_image(imageOne)

ws['A20'] = stringTwo
imageTwo = openpyxl.drawing.image.Image(r'C:\Users\VD102541\Desktop\PDX_DeArchive_Barcode.png')
imageTwo.anchor = 'E20'
ws.add_image(imageTwo)

ws['A40'] = stringThree
imageThree = openpyxl.drawing.image.Image(r"C:\Users\VD102541\Desktop\PDX_ArchiveRecycle_Barcode.png")
imageThree.anchor = 'E40'
ws.add_image(imageThree)

wb.save(r"C:\Users\VD102541\Desktop\barcodes.xlsx")
